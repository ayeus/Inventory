import openpyxl
import os
import logging

logger = logging.getLogger(__name__)

# File to read and update the inventory data
EXCEL_FILE = os.path.join(os.path.dirname(__file__), '../data/inventory_data.xlsx')

def load_excel_data():
    try:
        logger.debug("Loading Excel file from: %s", EXCEL_FILE)
        if not os.path.exists(EXCEL_FILE):
            logger.error("Excel file not found at: %s", EXCEL_FILE)
            return {}
        
        workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        inventory_data = {}
        
        for sheet_name in workbook.sheetnames:
            logger.debug("Processing sheet: %s", sheet_name)
            sheet = workbook[sheet_name]
            
            if not sheet.dimensions or sheet.dimensions == 'A1:A1':
                logger.warning("Sheet %s appears to be empty (dimensions: %s)", sheet_name, sheet.dimensions)
                has_data = False
                for row in sheet.rows:
                    for cell in row:
                        if cell.value is not None:
                            has_data = True
                            break
                    if has_data:
                        break
                if not has_data:
                    inventory_data[sheet_name] = []
                    continue
            
            min_row = 1
            max_row = max(sheet.max_row, 1)
            min_col = 1
            max_col = max(sheet.max_column, 1)
            
            for row in sheet.rows:
                for cell in row:
                    if cell.value is not None:
                        max_row = max(max_row, cell.row)
                        max_col = max(max_col, cell.column)
            
            logger.debug("Sheet %s dimensions: min_row=%d, max_row=%d, min_col=%d, max_col=%d", 
                        sheet_name, min_row, max_row, min_col, max_col)
            
            grid = []
            for row_idx in range(min_row, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value if cell.value is not None else ''
                    row_data.append(value)
                if any(cell != '' for cell in row_data):
                    grid.append(row_data)
            
            logger.debug("Grid for sheet %s: %s", sheet_name, grid)
            inventory_data[sheet_name] = grid
        
        return inventory_data
    except Exception as e:
        logger.error("Error loading Excel data: %s", str(e))
        return {}

def update_inventory_data(category, headers, new_entry, row_index=None):
    try:
        logger.debug("Updating inventory for category: %s, row_index: %s", category, row_index)
        logger.debug("Headers: %s", headers)
        logger.debug("New entry: %s", new_entry)

        data_dir = os.path.dirname(EXCEL_FILE)
        if not os.path.exists(data_dir):
            logger.debug("Creating data directory: %s", data_dir)
            os.makedirs(data_dir)

        if not os.path.exists(EXCEL_FILE):
            logger.error("Excel file not found at: %s", EXCEL_FILE)
            return False, "Excel file not found."

        logger.debug("Loading workbook: %s", EXCEL_FILE)
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        if category not in workbook.sheetnames:
            logger.debug("Creating new sheet for category: %s", category)
            sheet = workbook.create_sheet(title=category)
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
        else:
            sheet = workbook[category]
            logger.debug("Sheet %s already exists, max_row: %d", category, sheet.max_row)

        s_no_idx = None
        item_desc_idx = None
        stock_idx = None
        for col_idx, header in enumerate(headers):
            header_lower = str(header).lower()
            if 's.no' in header_lower or 'sno' in header_lower or 'sl. no' in header_lower:
                s_no_idx = col_idx
            if 'item description' in header_lower:
                item_desc_idx = col_idx
            if 'quantity' in header_lower or 'list' in header_lower or 'stock' in header_lower:
                stock_idx = col_idx

        if s_no_idx is None or item_desc_idx is None or stock_idx is None:
            logger.error("Required columns (S.No., Item Description, Stock) not found in headers: %s", headers)
            return False, "Required columns (S.No., Item Description, Stock) not found."

        logger.debug("Column indices - S.No.: %d, Item Description: %d, Stock: %d", s_no_idx, item_desc_idx, stock_idx)

        new_s_no = new_entry[s_no_idx]
        new_item_desc = new_entry[item_desc_idx]
        try:
            new_stock = int(new_entry[stock_idx]) if new_entry[stock_idx] and str(new_entry[stock_idx]).isdigit() else 0
        except ValueError:
            logger.error("Invalid stock value: %s", new_entry[stock_idx])
            return False, "Invalid stock value. Please enter a valid number."

        if row_index is not None:
            # Edit mode: Update the specific row
            if row_index < 2 or row_index > sheet.max_row:
                logger.error("Invalid row index: %d", row_index)
                return False, "Invalid row index."
            
            for col_idx, value in enumerate(new_entry, start=1):
                cell = sheet.cell(row=row_index, column=col_idx)
                cell.value = value
            logger.debug("Updated row %d with entry: %s", row_index, new_entry)
        else:
            # Add mode: Check if the item exists, update or append
            item_found = False
            for row_idx in range(2, sheet.max_row + 1):
                current_s_no = sheet.cell(row=row_idx, column=s_no_idx + 1).value
                current_item_desc = sheet.cell(row=row_idx, column=item_desc_idx + 1).value
                logger.debug("Checking row %d: S.No.=%s, Item Description=%s", row_idx, current_s_no, current_item_desc)
                if str(current_s_no) == str(new_s_no) or str(current_item_desc).lower() == str(new_item_desc).lower():
                    current_stock = sheet.cell(row=row_idx, column=stock_idx + 1).value
                    current_stock = int(current_stock) if current_stock and str(current_stock).isdigit() else 0
                    updated_stock = current_stock + new_stock
                    sheet.cell(row=row_idx, column=stock_idx + 1).value = updated_stock
                    item_found = True
                    logger.debug("Updated stock for item %s (S.No. %s): %d", new_item_desc, new_s_no, updated_stock)
                    break

            if not item_found:
                next_row = sheet.max_row + 1 if sheet.max_row > 1 else 2
                for col_idx, value in enumerate(new_entry, start=1):
                    cell = sheet.cell(row=next_row, column=col_idx)
                    cell.value = value
                logger.debug("Added new item %s (S.No. %s) with stock %d at row %d", new_item_desc, new_s_no, new_stock, next_row)

        logger.debug("Saving workbook to %s", EXCEL_FILE)
        workbook.save(EXCEL_FILE)
        logger.debug("Updated inventory in %s, sheet %s", EXCEL_FILE, category)
        return True, "Stock updated successfully."
    except Exception as e:
        logger.error("Error updating inventory: %s", str(e))
        return False, f"Error updating inventory: {str(e)}"

def delete_category_data(category):
    try:
        logger.debug("Deleting all data for category: %s", category)
        
        if not os.path.exists(EXCEL_FILE):
            logger.error("Excel file not found at: %s", EXCEL_FILE)
            return False, "Excel file not found."

        logger.debug("Loading workbook: %s", EXCEL_FILE)
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        if category not in workbook.sheetnames:
            logger.warning("Category %s not found in workbook", category)
            return False, f"Category '{category}' not found."

        # Delete the sheet
        workbook.remove(workbook[category])
        logger.debug("Removed sheet for category: %s", category)
        
        # Save the workbook
        workbook.save(EXCEL_FILE)
        logger.debug("Deleted category %s from %s", category, EXCEL_FILE)
        return True, "Category deleted successfully."
    except Exception as e:
        logger.error("Error deleting category: %s", str(e))
        return False, f"Error deleting category: {str(e)}"